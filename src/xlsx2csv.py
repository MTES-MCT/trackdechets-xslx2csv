import io
from pathlib import Path

import click
from InquirerPy import inquirer
from InquirerPy.base.control import Choice
from openpyxl import load_workbook
from rich.console import Console, Group
from rich.padding import Padding
from rich.panel import Panel

from core.constants import ETABLISSMENTS_FIELDS, ROLE_FIELDS
from core.models import EtabRows, RoleRows

console = Console()


def load_xlsx(file):
    wb = load_workbook(filename=file)
    sheetnames = wb.sheetnames
    if sheetnames[0] != "etablissements":
        console.print("Missing etablissements tab", style="bold red")
    if sheetnames[1] != "roles":
        console.print("Missing role tab", style="bold red")
    return wb


help_txt = """
[bold yellow]This cli tools parses, validates and converts bulk import xlsx files to ready to import csv.[/bold yellow]

[bold cyan]Usage:[/bold cyan]

    python xlsx2csv.py

    Expects a file name file.xlsx at the same level of the script

    python xlsx2csv.py path/to/my-file.xslx


[bold cyan]Operations:[/bold cyan]

    Tries to parse the xlsx file (xls are not accepted) and load it in memory and performs the following cleanups
        - remove unwanted extra surrounding spaces
        - remove blank spaces or dots sometimes inserted in sirets for readability
        - add 0 to phone numbers if missing due to excel number formatting
        - format phones
        - format companyTypes to be ingested by the Trackdéchets script
        - remove funky space chars
        - uppercase roles
        - lowercase email

    Performs the following validations:
        - file has the correct tabs
        - sirets have the correct length
        - sirets from role tabs are in etablissements tabs
        - emails are valid (no accents)
        - phones are valid (national format, with spaces, -, / or nothing) eg. 06-12-34-56-78
        - company types are valid
        - roles are valid
        - sirets from etablissemnt tabs have at jeast one ADMIN from roles tab
        - :backhand_index_pointing_right: verification in sirene database is not performed
    :thumbs_up: [red]If validation fails[/red], menu allows you to generate a french digest to send back to user who sent the file iot help him to correct the mistakes.
    :thumbs_down: [green]If validation successes[/green], menu allows csv files generation, ready to be imported into Trackdéchets
"""

CONTEXT_SETTINGS = dict(help_option_names=["-h", "--help"])


class InvalidHeaderException(Exception):
    pass


def validate_header(first_row, expected_header):
    """Validate first worksheet row matches `expected_header`"""

    header = [cell.value for cell in first_row]
    if header != expected_header:
        raise InvalidHeaderException


class RichHelp(click.Command):
    """Custom help formatter"""

    def format_help(self, ctx, formatter):
        sio = io.StringIO()
        h_console = Console(file=sio, force_terminal=True)
        h_console.print(help_txt)
        formatter.write(sio.getvalue())


@click.command(cls=RichHelp, context_settings=CONTEXT_SETTINGS)
@click.argument("file", default="file.xlsx")
def parse_validate_convert(file):
    wb = load_xlsx(file)

    ws_etablissements = wb.worksheets[0]
    ws_roles = wb.worksheets[1]

    etab_first_row = ws_etablissements[1][: len(ETABLISSMENTS_FIELDS)]
    role_first_row = ws_roles[1][: len(ROLE_FIELDS)]

    try:
        validate_header(etab_first_row, ETABLISSMENTS_FIELDS)

    except InvalidHeaderException:

        console.print(
            ":thumbs_down: [red]Etablissement header does not match expected cells"
        )
        return

    try:

        validate_header(role_first_row, ROLE_FIELDS)
    except InvalidHeaderException:

        console.print(":thumbs_down: [red]Role header does not match expected cells")
        return

    etab_rows = EtabRows.from_worksheet(ws_etablissements)

    etab_rows.validate()

    role_rows = RoleRows.from_worksheet(ws_roles)

    role_rows.validate(etab_rows.sirets())

    # This validation can occur when both tabs are already validated
    if etab_rows.is_valid and role_rows.is_valid:
        etab_rows.validate_have_admin(role_rows.admin_sirets())

    if etab_rows.is_valid:
        console.print(":thumbs_up: [green bold]Etab tab is valid")
    else:
        console.print(":thumbs_down: [red]Etablissement tab has errors")
    if role_rows.is_valid:
        console.print(":thumbs_up: [green bold]Role tab is valid")
    else:
        console.print(":thumbs_down: [red]Role tab has errors")

    menu_choices = []

    if etab_rows.is_valid and role_rows.is_valid:

        menu_choices.append(Choice(value="view_etabs", name="View etablissements"))
        menu_choices.append(Choice(value="view_roles", name="View roles"))
        menu_choices.append(Choice(value="export", name="Export csv"))
    else:
        menu_choices.append(Choice(value="view_etabs", name="View etablissements"))

        menu_choices.append(Choice(value="message", name="Generate message"))
    menu_choices.append(Choice(value=None, name="Exit"))

    action = inquirer.select(
        message="Select an action:",
        choices=menu_choices,
        default=None,
    ).execute()

    if action == "message":
        etab_rows.make_messages()
        role_rows.make_messages()

    if action == "view_etabs":
        etab_rows.as_table()
    if action == "view_roles":
        role_rows.as_table()

    if action == "export":
        console.print(":weight_lifter:[green] Exporting")
        csv_dir = Path(".") / "csv"
        csv_dir.mkdir(exist_ok=True)
        with open("csv/etablissements.csv", "w", newline="") as csvfile:
            for row in etab_rows.as_csv():
                csvfile.write(row)
        with open("csv/roles.csv", "w", newline="") as csvfile:
            for row in role_rows.as_csv():
                csvfile.write(row)

        console.print(":thumbs_up:[green] Done")

        panel_group = Group(
            Padding(
                "[green] scalingo --app trackdechets-production-api --region osc-secnum-fr1 run --file ./csv bash",
                (1, 4),
            ),
            Padding("[yellow] then : ", (1, 0)),
            Padding("[green] tar -C /tmp -xvf /tmp/uploads/csv.tar.gz", (1, 4)),
            Padding("[and] and : ", (1, 0)),
            Padding(
                "[green] tsx --tsconfig back/tsconfig.lib.json back/src/users/bulk-creation/index.ts --validateOnly --csvDir=/tmp/",
                (1, 4),
            ),
            Padding("[or] and : ", (1, 0)),
            Padding(
                "[green] tsx --tsconfig back/tsconfig.lib.json back/src/users/bulk-creation/index.ts --csvDir=/tmp/",
                (1, 4),
            ),
        )
        console.print(
            Panel(
                panel_group,
                title="[bold green]Now you may run these commands to import csvs",
            )
        )

        # console.print(Panel("Hello, [red]World!"))
    console.print(":wave:[green] Exiting")


if __name__ == "__main__":
    parse_validate_convert()
